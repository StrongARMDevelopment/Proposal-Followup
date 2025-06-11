import os
import pandas as pd
import datetime
import logging
import win32com.client as win32
from openpyxl import load_workbook
import time
import configparser
import sys
import zipfile
import ast
import shutil

# --- Global Script Constants ---
CONFIG_FILE_NAME = "config.ini"
SCRIPT_VERSION = "2.2.0" # Version updated for new features
EXCEL_DATE_OFFSET = datetime.datetime(1899, 12, 30)
EXCEL_ROW_OFFSET = 2  # Excel rows are 1-based, plus header

# --- Global Variables ---
# These are managed within the main execution block and passed as needed.
outlook = None
namespace = None
outlook_signature = ""
lock_file_handle = None

# --- Helper & Setup Functions ---

def excel_date_to_datetime(excel_date):
    """Convert Excel serial date to datetime."""
    if isinstance(excel_date, (int, float)) and excel_date > 0:
        return EXCEL_DATE_OFFSET + datetime.timedelta(days=excel_date)
    return pd.to_datetime(excel_date, errors='coerce')

def create_lock_file(lock_file_path):
    """Create a lock file to prevent concurrent script runs."""
    global lock_file_handle
    try:
        lock_file_handle = open(lock_file_path, 'x')
        logging.info(f"Lock file created: {lock_file_path}")
        return True
    except FileExistsError:
        logging.error(f"Lock file {lock_file_path} already exists. Another instance may be running.")
        return False
    except IOError as e:
        logging.error(f"Failed to create lock file {lock_file_path}: {e}")
        return False

def remove_lock_file(lock_file_path):
    """Remove the lock file on script exit."""
    global lock_file_handle
    try:
        if lock_file_handle:
            lock_file_handle.close()
            lock_file_handle = None
        if os.path.exists(lock_file_path):
            os.remove(lock_file_path)
            logging.info(f"Lock file removed: {lock_file_path}")
    except IOError as e:
        logging.error(f"Failed to remove lock file {lock_file_path}: {e}")

def load_configuration():
    """Load and parse the configuration file."""
    parser = configparser.ConfigParser()
    if not os.path.exists(CONFIG_FILE_NAME):
        logging.critical(f"CRITICAL: Configuration file '{CONFIG_FILE_NAME}' not found. Exiting.")
        raise FileNotFoundError(f"Configuration file '{CONFIG_FILE_NAME}' not found.")
    try:
        parser.read(CONFIG_FILE_NAME)
        # Pre-process list-like settings for easier access later
        parser.set('Settings', 'YearsToProcess', str([year.strip() for year in parser.get('Settings', 'YearsToProcess', fallback='').split(',') if year.strip()]))
        parser.set('Settings', 'ValidMonths', str([month.strip() for month in parser.get('Settings', 'ValidMonths', fallback='').split(',') if month.strip()]))
        logging.info(f"Configuration loaded from '{CONFIG_FILE_NAME}'.")
        return parser
    except configparser.Error as e:
        logging.critical(f"CRITICAL: Error parsing configuration file '{CONFIG_FILE_NAME}': {e}. Exiting.")
        raise

def initialize_outlook(config):
    """Initialize Outlook application and retrieve signature."""
    global outlook, namespace, outlook_signature
    # In test email mode, we need Outlook, so we don't check for DryRun here.
    # The main() block will handle when to call this.
    try:
        outlook = win32.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        logging.info("Outlook MAPI session initialized.")

        try:
            dummy_mail = outlook.CreateItem(0)
            if not dummy_mail.HTMLBody:
                dummy_mail.Display()
                time.sleep(0.5)
            outlook_signature = dummy_mail.HTMLBody
            dummy_mail.Close(0)
            if outlook_signature:
                logging.info("Successfully retrieved Outlook signature.")
            else:
                logging.warning("Outlook signature is empty or could not be retrieved.")
        except Exception as e:
            logging.error(f"Could not retrieve Outlook signature: {e}.")
            outlook_signature = ""
    except Exception as e:
        logging.error(f"Outlook is not available or not configured properly: {e}. Emails will not be sent.")
        outlook = None
    return outlook # Return the outlook object for use in main()

def get_outlook_account(outlook_app, desired_email):
    """Get the Outlook account matching the desired email."""
    if not outlook_app or not desired_email:
        return None
    try:
        for acc in outlook_app.Session.Accounts:
            if acc.SmtpAddress.lower() == desired_email.lower():
                logging.info(f"Using specified Outlook account: {desired_email}")
                return acc
        logging.warning(f"Specified Outlook account '{desired_email}' not found. Using default.")
    except Exception as e:
        logging.error(f"Error trying to get specified Outlook account: {e}")
    return None

def setup_logging(config):
    """Set up logging based on config."""
    log_file_path = config.get('Paths', 'ScriptLogFile', fallback='followup_automation.log')
    log_level = getattr(logging, config.get('Settings', 'LogLevel', fallback='INFO').upper(), logging.INFO)
    
    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)
    
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)

    file_handler = logging.FileHandler(log_file_path, mode='w')
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    root_logger.addHandler(console_handler)

# --- Core Processing Functions ---

def process_proposals(config, today_date):
    """Process proposals from Excel files and group follow-ups by contact email."""
    grouped_follow_ups = {}
    cols = config['Columns']
    required_df_cols = [
        cols['DateProposalSubmitted'], cols['LastCorrespondence'], cols['ContactEmail'],
        cols['ContactName'], cols['ProjectName'], cols['Value'],
        cols['Won'], cols['Lost'], cols['ReBid'], cols['FollowUpStage']
    ]
    years_to_process = ast.literal_eval(config.get('Settings', 'YearsToProcess'))
    valid_months_list = ast.literal_eval(config.get('Settings', 'ValidMonths'))

    for year in years_to_process:
        try:
            file_path = config.get('Paths', f'Log{year}Path')
        except configparser.NoOptionError:
            logging.warning(f"No path configured for year {year} in '{CONFIG_FILE_NAME}'. Skipping.")
            continue

        if not os.path.exists(file_path):
            logging.error(f"File not found: {file_path} for year {year}. Skipping this file.")
            continue

        logging.info(f"Processing file: {file_path} for year {year}")
        try:
            df_sheets = pd.read_excel(file_path, sheet_name=None, engine="openpyxl", usecols=lambda x: x in required_df_cols)
        except (FileNotFoundError, PermissionError, zipfile.BadZipFile, ValueError) as e:
            logging.error(f"Error reading Excel file {file_path}: {e}. Skipping.")
            continue
        except Exception as e:
            logging.error(f"Could not read Excel file {file_path} using pandas: {e}. Skipping this file.")
            continue

        for sheet_name, df in df_sheets.items():
            if sheet_name not in valid_months_list:
                continue

            logging.info(f"Collecting data from {year} - Sheet: {sheet_name}")

            # Validate columns are present
            missing_in_df = [col for col in required_df_cols if col not in df.columns]
            if missing_in_df:
                logging.error(f"Required columns {missing_in_df} not found in sheet '{sheet_name}'. Skipping sheet.")
                continue

            if cols['FollowUpStage'] not in df.columns:
                df[cols['FollowUpStage']] = 0

            for index, row in df.iterrows():
                try:
                    submission_date_val = row[cols['DateProposalSubmitted']]
                    submission_date = excel_date_to_datetime(submission_date_val)
                    if pd.isna(submission_date) or submission_date.year < 2000:
                        continue
                    submission_date_dt_date = submission_date.date()
                    
                    contact_email = str(row.get(cols['ContactEmail'], "") or "").strip().lower()
                    if not contact_email or "@" not in contact_email:
                        continue

                    if str(row.get(cols['Won'], "")).strip().upper() == "X" or \
                       str(row.get(cols['Lost'], "")).strip().upper() == "X" or \
                       str(row.get(cols['ReBid'], "")).strip().upper() == "X":
                        continue

                    follow_up_stage = int(row.get(cols['FollowUpStage'], 0))
                    last_corr_val = row[cols['LastCorrespondence']]
                    last_corr_date = excel_date_to_datetime(last_corr_val).date() if pd.notna(last_corr_val) else None
                    
                    template_index, ready_for_follow_up = -1, False
                    days_first = config.getint('Settings', 'DaysFirstFollowUp')
                    days_subsequent = config.getint('Settings', 'DaysSubsequentFollowUps')

                    if follow_up_stage == 0:
                        relevant_date = last_corr_date or submission_date_dt_date
                        if (today_date - relevant_date).days >= days_first:
                            template_index, ready_for_follow_up = 0, True
                    elif follow_up_stage > 0 and last_corr_date:
                        if (today_date - last_corr_date).days >= days_subsequent:
                            template_index, ready_for_follow_up = (1 if follow_up_stage == 1 else 2), True
                    
                    if not ready_for_follow_up:
                        continue

                    val = row[cols['Value']]
                    value_str = f"{float(val):,.2f}" if pd.notna(val) and str(val).strip() != "" else ""
                    
                    proposal_details = {
                        "project_name": str(row.get(cols['ProjectName'], 'N/A')),
                        "submission_date_str": submission_date.strftime('%m-%d-%Y'),
                        "value_str": value_str,
                        "contact_first_name": str(row.get(cols['ContactName'], "")).strip().split()[0] or "there",
                        "row_index_df": index, "sheet_name": sheet_name, "workbook_path": file_path,
                        "new_follow_up_stage": follow_up_stage + 1, "email_snippet_key": template_index,
                        "last_correspondence_col_df_idx": df.columns.get_loc(cols['LastCorrespondence']),
                        "follow_up_stage_col_df_idx": df.columns.get_loc(cols['FollowUpStage'])
                    }
                    grouped_follow_ups.setdefault(contact_email, []).append(proposal_details)
                    logging.info(f"Queued follow-up for '{proposal_details['project_name']}' to {contact_email} (Excel Row: {index+EXCEL_ROW_OFFSET}).")

                except Exception as e:
                    logging.error(f"Error on row {index+EXCEL_ROW_OFFSET} in {sheet_name}: {e}", exc_info=True)

    logging.info(f"--- Phase 1 Complete: Collected {sum(len(v) for v in grouped_follow_ups.values())} total qualifying follow-ups. ---")
    return grouped_follow_ups

def send_consolidated_emails(config, grouped_follow_ups, outlook_app, is_dry_run=False):
    """Send consolidated emails for grouped follow-ups."""
    global outlook_signature
    successful_sends = []
    if not grouped_follow_ups:
        logging.info("No qualifying follow-ups to send. Skipping Email Phase.")
        return successful_sends
    if not outlook_app and not is_dry_run:
        logging.error("Outlook not initialized. Skipping Email Phase.")
        return successful_sends

    logging.info(f"--- Phase 2: Sending Consolidated Emails {'(DRY RUN)' if is_dry_run else ''} ---")
    
    email_cfg, subject_cfg, snippet_cfg = config['EmailBody'], config['EmailSubjects'], config['EmailSnippets']

    for contact_email, proposals in grouped_follow_ups.items():
        project_names = [p['project_name'] for p in proposals]
        subject_line = subject_cfg.get(
            'SingleProject' if len(project_names) == 1 else 'TwoProjects' if len(project_names) == 2 else 'MultipleProjects',
            'Follow-up on proposal(s)'
        ).format(project_name=project_names[0], project_name_1=project_names[0], project_name_2=project_names[1] if len(project_names)>1 else '', first_project_name=project_names[0])

        body_html = email_cfg.get('Greeting').format(contact_first_name=proposals[0]['contact_first_name'])
        body_html += email_cfg.get('Intro')
        body_html += email_cfg.get('ProjectListStart')
        for p in proposals:
            snippet = snippet_cfg.get(f"Snippet_{p['email_snippet_key']}", "Status update requested.")
            value_info = email_cfg.get('ValueFormat').format(value_str=p['value_str']) if p['value_str'] else ""
            body_html += email_cfg.get('ProjectListItem').format(
                project_name=p['project_name'], submission_date_str=p['submission_date_str'],
                value_info=value_info, snippet_text=snippet
            )
        body_html += email_cfg.get('ProjectListEnd')
        body_html += email_cfg.get('Closing')

        # Append signoff and signature
        if outlook_signature:
            body_html += outlook_signature
        else:
            body_html += email_cfg.get('Signoff', '<p>Best regards,</p>')
            body_html += email_cfg.get('DefaultSignatureFallback', '<br>Your Name')

        if is_dry_run:
            logging.info(f"DRY RUN: Would send to {contact_email} - Subject: '{subject_line}'")
            logging.debug(f"DRY RUN: Email body for {contact_email}:\n{body_html}")
            successful_sends.extend(proposals)
            continue

        sent_successfully = False
        for attempt in range(config.getint('Settings', 'MaxEmailSendAttempts', fallback=3)):
            try:
                mail = outlook_app.CreateItem(0)
                account = get_outlook_account(outlook_app, config.get('Settings', 'DesiredOutlookAccount'))
                if account: mail.SendUsingAccount = account
                mail.To, mail.Subject, mail.HTMLBody = contact_email, subject_line, body_html
                mail.Send()
                logging.info(f"Email sent to {contact_email} for {len(proposals)} project(s).")
                sent_successfully = True
                break
            except Exception as e:
                logging.warning(f"Attempt {attempt+1} failed to send to {contact_email}: {e}")
                time.sleep(config.getint('Settings', 'EmailRetryDelaySeconds', fallback=5))
        
        if sent_successfully:
            successful_sends.extend(proposals)
            time.sleep(config.getint('Settings', 'EmailSendDelaySeconds', fallback=1))
        else:
            logging.error(f"Failed to send email to {contact_email} after all attempts.")

    logging.info(f"--- Phase 2 Complete. ---")
    return successful_sends

def update_excel_sheets(config, successful_updates, today_date):
    """Update Excel sheets for proposals that were successfully emailed."""
    is_dry_run = config.getboolean('Settings', 'DryRun')
    if not successful_updates or is_dry_run:
        log_msg = "No updates to save." if not successful_updates else "DRY RUN: Skipping Excel updates."
        logging.info(f"--- Phase 3: {log_msg} ---")
        if is_dry_run and successful_updates: # Log what would have been updated
             for item in successful_updates:
                 logging.info(f"DRY RUN: Would update {item['workbook_path']}, Sheet '{item['sheet_name']}', Row {item['row_index_df']+EXCEL_ROW_OFFSET} "
                              f"(Project: '{item['project_name']}') to Stage {item['new_follow_up_stage']}.")
        return

    logging.info(f"--- Phase 3: Updating Excel Sheets ---")
    for wb_path, proposals in pd.DataFrame(successful_updates).groupby('workbook_path'):
        try:
            if config.getboolean('Settings', 'BackupExcelBeforeSave', fallback=True):
                shutil.copy2(wb_path, wb_path + ".bak")
                logging.info(f"Backup created for {wb_path}")

            wb = load_workbook(wb_path)
            for sheet_name, sheet_updates in proposals.groupby('sheet_name'):
                if sheet_name not in wb.sheetnames:
                    logging.error(f"Sheet '{sheet_name}' not in {wb_path}. Skipping.")
                    continue
                ws = wb[sheet_name]
                logging.info(f"Updating sheet: '{sheet_name}' in '{wb_path}'.")
                for item in sheet_updates.to_dict('records'):
                    row, lc_col, st_col = item['row_index_df'] + EXCEL_ROW_OFFSET, item['last_correspondence_col_df_idx'] + 1, item['follow_up_stage_col_df_idx'] + 1
                    ws.cell(row=row, column=lc_col, value=today_date.strftime('%m-%d-%Y'))
                    ws.cell(row=row, column=st_col, value=item['new_follow_up_stage'])
            wb.save(wb_path)
            logging.info(f"Successfully saved changes to {wb_path}")
        except Exception as e:
            logging.error(f"Failed to update {wb_path}: {e}", exc_info=True)
    logging.info(f"--- Phase 3 Complete. ---")

def send_comprehensive_test_emails(config, outlook_app):
    """
    Sends a test email for each template variation (single, two, multiple projects)
    to the TestEmailRecipient specified in config.ini.
    """
    test_recipient = config.get('Settings', 'TestEmailRecipient', fallback=None)
    if not test_recipient:
        logging.error("No TestEmailRecipient specified in config. Cannot send test emails.")
        return

    today_str = datetime.datetime.today().strftime('%m-%d-%Y')

    # 1. Single project email
    single_project = [
        {
            "project_name": "Test Project Alpha",
            "submission_date_str": today_str,
            "value_str": "10,000.00",
            "contact_first_name": "Test",
            "row_index_df": 0,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 0,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        }
    ]
    # 2. Two projects email
    two_projects = [
        {
            "project_name": "Test Project Beta",
            "submission_date_str": today_str,
            "value_str": "20,000.00",
            "contact_first_name": "Test",
            "row_index_df": 1,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 1,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        },
        {
            "project_name": "Test Project Gamma",
            "submission_date_str": today_str,
            "value_str": "30,000.00",
            "contact_first_name": "Test",
            "row_index_df": 2,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 2,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        }
    ]
    # 3. Multiple projects email (3+)
    multiple_projects = [
        {
            "project_name": "Test Project Delta",
            "submission_date_str": today_str,
            "value_str": "40,000.00",
            "contact_first_name": "Test",
            "row_index_df": 3,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 0,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        },
        {
            "project_name": "Test Project Epsilon",
            "submission_date_str": today_str,
            "value_str": "50,000.00",
            "contact_first_name": "Test",
            "row_index_df": 4,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 1,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        },
        {
            "project_name": "Test Project Zeta",
            "submission_date_str": today_str,
            "value_str": "60,000.00",
            "contact_first_name": "Test",
            "row_index_df": 5,
            "sheet_name": "TestSheet",
            "workbook_path": "Test.xlsx",
            "new_follow_up_stage": 1,
            "email_snippet_key": 2,
            "last_correspondence_col_df_idx": 0,
            "follow_up_stage_col_df_idx": 0
        }
    ]

    # Send each test email as a separate message
    for test_case, proposals in [
        ("Single Project", single_project),
        ("Two Projects", two_projects),
        ("Multiple Projects", multiple_projects)
    ]:
        test_data = {test_recipient: proposals}
        logging.info(f"Sending comprehensive test email: {test_case}")
        send_consolidated_emails(config, test_data, outlook_app, is_dry_run=False)


# --- Main Execution ---

def main():
    """Main script execution."""
    today_date = datetime.datetime.today().date()
    
    # Minimal initial logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    
    try:
        config = load_configuration()
    except Exception:
        sys.exit(1) # Error already logged

    setup_logging(config)
    logging.info(f"--- Script Started (Version: {SCRIPT_VERSION}) ---")
    
    lock_file_path = config.get('Paths', 'LockFilePath', fallback='followup_automation.lock')
    if not create_lock_file(lock_file_path):
        sys.exit(1)

    outlook_app = None # Initialize to None
    try:
        is_dry_run = config.getboolean('Settings', 'DryRun')
        is_test_mode = config.getboolean('Settings', 'SendTestEmail', fallback=False)

        if is_test_mode:
            outlook_app = initialize_outlook(config)
            if outlook_app:
                send_comprehensive_test_emails(config, outlook_app)
            else:
                logging.error("Outlook not available. Cannot send test email.")
        elif is_dry_run:
             logging.warning("IMPORTANT: Script is running in DRY RUN mode. No emails or Excel changes will be made.")
             collected = process_proposals(config, today_date)
             update_excel_sheets(config, collected, today_date) # Pass `collected` to log what would be updated
        else: # Live Run
            outlook_app = initialize_outlook(config)
            if outlook_app:
                collected = process_proposals(config, today_date)
                sent = send_consolidated_emails(config, collected, outlook_app)
                update_excel_sheets(config, sent, today_date)
            else:
                logging.error("Outlook not available. Halting live run.")

    except Exception as e:
        logging.critical(f"An unhandled error occurred in main execution: {e}", exc_info=True)
    finally:
        remove_lock_file(lock_file_path)
        logging.info("--- Script Finished ---")

if __name__ == "__main__":
    main()